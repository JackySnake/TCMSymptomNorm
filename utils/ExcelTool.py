# -*-coding:utf-8 -*-
import openpyxl


class ExcelReader(object):
    def __init__(self, path):
        self.__wbk = openpyxl.load_workbook(path, read_only=True)

    def getSheetName(self):
        return self.__wbk.sheetnames

    def getSheetData(self, st_name):
        st = self.__wbk[st_name]
        return [[str(cell.value).strip() for cell in row] for row in st.rows]

    def getAllData(self):
        return {st_name: self.getSheetData(st_name) for st_name in self.getSheetName()}


class ExcelWriter(object):
    def __init__(self, path, mode):
        if mode == 'a':
            self.__wbk = openpyxl.load_workbook(path, read_only=False)
        elif mode == 'w':
            self.__wbk = openpyxl.Workbook()
        else:
            print(f'Error: {mode} is a error mode !!!!')
        self.__path = path

    def createSheet(self, st_name):
        return self.__wbk.create_sheet(st_name)

    def writeData(self, st_name, values):
        rows = [[i] * len(values[0]) for i in range(1, len(values) + 1)]
        cols = [list(range(1, len(values[0]) + 1))] * len(values)
        if st_name in self.__wbk.get_sheet_names():
            st = self.__wbk.get_sheet_by_name(st_name)
        else:
            st = self.createSheet(st_name)
        for value, row, col in zip(values, rows, cols):
            for v, r, c in zip(value, row, col):
                st.cell(row=r, column=c).value = v

    def save(self):
        self.__wbk.save(self.__path)

    def deleteSheet(self, st_name):
        if st_name in self.__wbk.get_sheet_names():
            self.__wbk.remove_sheet(self.__wbk.get_sheet_by_name(st_name))


# if __name__ == '__main__':
#     Excel = ExcelReader()
#     print(Excel.getSheetData('气'))
